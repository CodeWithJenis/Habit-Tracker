import pandas as pd
from matplotlib import pyplot as plt
from datetime import date
import json
from openpyxl.drawing.image import Image as XLImage
import os


def save_pie_chart(values, labels, title, path, colors):
    """Reusable function to create and save a pie chart image."""
    plt.figure(figsize=(4, 3))
    plt.pie(
        values,
        labels=labels,
        autopct="%.2f%%",
        colors=colors,
        explode=(0.05,) * len(values),
        shadow=True,
        wedgeprops={'edgecolor': 'black'}
    )
    plt.title(title, fontsize=14)
    plt.tight_layout()
    plt.savefig(path)
    plt.close()


def input_goals():
    """Prompt the user to input goals and their associated activities."""
    goals = []
    ng = int(input("How many goals do you have? "))
    
    for i in range(ng):
        goal_name = input(f"Enter Goal {i+1}: ")
        
        # Collect activities for this goal
        do_activities = []
        do_not_activities = []
        
        do_count = int(input(f"How many 'do' activities for '{goal_name}'? "))
        for j in range(do_count):
            activity = input(f"Enter 'do' activity {j+1} for '{goal_name}': ")
            do_activities.append(activity)

        do_not_count = int(input(f"How many 'do not' activities for '{goal_name}'? "))
        for j in range(do_not_count):
            activity = input(f"Enter 'do not' activity {j+1} for '{goal_name}': ")
            do_not_activities.append(activity)
        
        # Store in a structured way
        goals.append({
            "name": goal_name,
            "do": do_activities,
            "do_not": do_not_activities
        })
    
    return goals


def display_goals(goals):
    """Display all entered goals and their activities."""
    for i, goal in enumerate(goals):
        print(f"\nGoal {i+1}: {goal['name']}")
        print("  Do Activities:")
        for activity in goal["do"]:
            print(f"    - {activity}")
        print("  Do Not Activities:")
        for activity in goal["do_not"]:
            print(f"    - {activity}")

def tracker_goals(goals):
    """Track daily progress for each goal and activity."""
    today = date.today().isoformat()
    dates = input(f"Press Enter for {today} or Enter previous dates (YYYY-MM-DD):") or today
    TICK = "✔️"
    CROSS = "❌"
    goals_progress = {
        'date': dates,
        'goals': []
    }
    for i, goal in enumerate(goals):
        goal_result = {
            'name': goal["name"],
            'do': [],
            'do_not': []
        }
        print(f"\nGoal {i+1}: {goal['name']}")
        for activity in goal["do"]:
            print(f"    - {activity}")
            status = input("Completed (y/n?): ").strip().lower()
            goal_result['do'].append({
                'activity': activity,
                'status': TICK if status == "y" else CROSS
            })

        print("  Do Not Activities:")
        for activity in goal["do_not"]:
            print(f"    - {activity}")
            status = input("Did you AVOID it? (y/n):").strip().lower()
            goal_result['do_not'].append({
                'activity': activity,
                'status': TICK if status == "y" else CROSS
            })
        goals_progress['goals'].append(goal_result)
    print(json.dumps(goals_progress, indent=2, ensure_ascii=False))
    return goals_progress

def flatten_goals_progress(goals_progress):
    """Flatten the nested goals progress structure into lists for Excel export."""
    track_do = []
    track_do_not=[]
    date = goals_progress['date']

    for goal in goals_progress['goals']:
        # Process Do Activities
        for item in goal['do']:
            track_do.append({
                "Goal":goal['name'],
                "Date": date,
                "Do Activity": item['activity'],
                "Completed": item['status']
            })

        # Process Do Not Activities
        for item in goal['do_not']:
            track_do_not.append({
                "Goal":goal['name'],
                "Date": date,
                "Do Not Activity": item['activity'],
                "Avoided": item['status']
            })
    return track_do,track_do_not


def export_to_excel(goals, track_do, track_do_not):
    """Export goals and tracking data to an Excel file with formatted sheets."""
    temp_images = []  # List to keep track of all image paths for cleanup
    with pd.ExcelWriter(r"D:\Carrer\AI\AI Er Learning\Projects\Habit Tracker\goals.xlsx", engine="openpyxl") as writer:
        # --- Goal Sheet ---
        current_row = 0
        for goal in goals:
            # Prepare DataFrame for this goal
            max_len = max(len(goal['do']), len(goal['do_not']))
            do_list = goal['do'] + [''] * (max_len - len(goal['do']))
            do_not_list = goal['do_not'] + [''] * (max_len - len(goal['do_not']))
            df_goal = pd.DataFrame({
                'Do Activity': do_list,
                'Do Not Activity': do_not_list
            })
            # Write goal name above the table
            df_goal.to_excel(writer, sheet_name='Goal', startrow=current_row + 1, index=False)
            worksheet = writer.sheets['Goal']
            worksheet.cell(row=current_row + 1, column=1, value=goal['name'])
            current_row += len(df_goal) + 3  # Space between tables

        # --- Tracker Sheet ---
        current_row = 0
        compare_goal=[]
        all_goals = list(dict.fromkeys(item['Goal'] for item in track_do))
        worksheet = writer.book['Tracker'] if 'Tracker' in writer.book.sheetnames else writer.book.create_sheet('Tracker')
        for goal in all_goals:
            df_do = pd.DataFrame([item for item in track_do if item['Goal'] == goal])
            df_do_not = pd.DataFrame([item for item in track_do_not if item['Goal'] == goal])

            # Align lengths for side-by-side columns
            max_len = max(len(df_do), len(df_do_not))
            df_do = df_do.reindex(range(max_len), fill_value='')
            df_do_not = df_do_not.reindex(range(max_len), fill_value='')

            df_tracker = pd.DataFrame({
                'Date': df_do['Date'],
                'Do Activity': df_do['Do Activity'],
                'Completed': df_do['Completed'],
                'Do Not Activity': df_do_not['Do Not Activity'],
                'Avoided': df_do_not['Avoided']
            })

            # Write goal name above the table
            df_tracker.to_excel(writer, sheet_name='Tracker', startrow=current_row + 1, index=False)
            worksheet.cell(row=current_row + 1, column=1, value=goal)

            # Pie charts using the helper function
            completed = (df_do['Completed'] == "✔️").sum()
            not_completed = (df_do['Completed'] == "❌").sum()
            avoided = (df_do_not['Avoided'] == "✔️").sum()
            not_avoided = (df_do_not['Avoided'] == "❌").sum()
            total_progress = completed + avoided
            missed_progress = not_completed + not_avoided

            # Do Activity Pie
            img_path_do = f"D:/Carrer/AI/AI Er Learning/Projects/Habit Tracker/{goal}_Do_Activity_Pie.png"
            save_pie_chart([completed, not_completed], ["Completed", "Not Completed"], f"{goal} Do Activity Progress", img_path_do, ['green', 'red'])
            img_do = XLImage(img_path_do)
            img_do.anchor = f"A{current_row + len(df_tracker) + 3}"
            worksheet.add_image(img_do)
            temp_images.append(img_path_do)

            # Do Not Activity Pie
            img_path_do_not = f"D:/Carrer/AI/AI Er Learning/Projects/Habit Tracker/{goal}_Do_Not_Activity_Pie.png"
            save_pie_chart([avoided, not_avoided], ["Avoided", "Not Avoided"], f"{goal} Do Not Activity Progress", img_path_do_not, ['green', 'red'])
            img_do_not = XLImage(img_path_do_not)
            img_do_not.anchor = f"H{current_row + len(df_tracker) + 3}"
            worksheet.add_image(img_do_not)
            temp_images.append(img_path_do_not)

            # Overall Progress Pie
            img_path_overall = f"D:/Carrer/AI/AI Er Learning/Projects/Habit Tracker/{goal}_Overall_Progress_Pie.png"
            save_pie_chart([total_progress, missed_progress], ["Total Progress", "Missed Progress"], f"{goal} Overall Progress", img_path_overall, ['green', 'red'])
            img_overall = XLImage(img_path_overall)
            img_overall.anchor = f"P{current_row + len(df_tracker) + 3}"
            worksheet.add_image(img_overall)
            temp_images.append(img_path_overall)

            current_row += max(len(df_tracker) + 3, 23)
            compare_goal.append({
                'name': goal,
                'total progress': total_progress,
                'missed progress': missed_progress
            })

        # Bar chart for comparison
        df_compare_goal = pd.DataFrame(compare_goal)
        ax = df_compare_goal.plot(
            kind='bar',
            x='name',
            rot=45,
            figsize=(4, 4),
            legend=True,
            color=['green', 'red']
        )
        plt.title('Goal Progress Comparison')
        plt.xlabel('Goal')
        plt.ylabel('Count')
        plt.tight_layout()
        bar_img_path = r"D:/Carrer/AI/AI Er Learning/Projects/Habit Tracker/Goal_Comparision_Bar.png"
        plt.savefig(bar_img_path)
        plt.close()
        bar_img = XLImage(bar_img_path)
        bar_img.anchor = f"C{current_row + 5}"
        worksheet.add_image(bar_img)
        temp_images.append(bar_img_path)

    # --- Remove all temporary images after Excel is saved ---
    for img_path in temp_images:
        try:
            if os.path.exists(img_path):
                os.remove(img_path)
        except Exception as e:
            print(f"Could not remove {img_path}: {e}")

def main():
    """Main function to run the habit tracker workflow."""
    goals = input_goals()
    display_goals(goals)
    goals_progress = tracker_goals(goals)
    track_do, track_do_not = flatten_goals_progress(goals_progress)
    export_to_excel(goals, track_do, track_do_not)


if __name__ == "__main__":
    main()
